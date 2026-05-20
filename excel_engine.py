import os
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
import utils


CURRENCY_FORMAT = '"$"#,##0;-"$"#,##0'
PERCENT_FORMAT = '0.00%'
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
TITLE_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN_BORDER = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)


def _safe_filename(value):
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "-", str(value or "SITFA")).strip("-")
    return cleaned or "SITFA"


def _rgb_hex(rgb):
    if isinstance(rgb, (tuple, list)) and len(rgb) >= 3:
        return "".join(f"{max(0, min(255, int(part))):02X}" for part in rgb[:3])
    return "D9E2F3"


def _currency_to_number(value):
    if isinstance(value, (int, float)):
        return value
    text = str(value or "").replace("$", "").replace(" ", "").replace("\u00a0", "")
    text = text.split(",")[0].replace(".", "")
    text = re.sub(r"[^\d-]", "", text)
    try:
        return int(text)
    except ValueError:
        return value


def _percent_to_number(value):
    if isinstance(value, (int, float)):
        return value
    text = str(value or "").strip().replace("%", "").replace(",", ".")
    try:
        return float(text) / 100
    except ValueError:
        return value


def _decimal_to_number(value):
    if isinstance(value, (int, float)):
        return value
    text = str(value or "").strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return value


def _set_cell(cell, value, header=""):
    header_lower = str(header).lower()
    text = str(value or "")

    if text.startswith("$"):
        cell.value = _currency_to_number(value)
        cell.number_format = CURRENCY_FORMAT
    elif "%" in text:
        cell.value = _percent_to_number(value)
        cell.number_format = PERCENT_FORMAT
    elif header_lower in ("monto utm",):
        cell.value = _decimal_to_number(value)
        cell.number_format = "0.00000"
    else:
        cell.value = value


def _style_range(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _autosize_columns(ws, max_width=55):
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            width = max(width, min(max_width, len(str(value)) + 2))
        ws.column_dimensions[letter].width = width


def _headers_for_reajuste(reajuste_tipo):
    if reajuste_tipo == "UTM":
        return ["Desde", "Hasta", "Meses", "Monto UTM", "Valor UTM", "Total Pesos"]
    if reajuste_tipo == "EMOLUMENTOS":
        return ["Desde", "Hasta", "Renta imponible", "Descuentos legales", "Base de calculo", "%", "Total"]
    return ["Desde", "Hasta", "Meses", "Reajuste", "Pension reajustada", "Total"]


def _write_title(ws, title, row):
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = TITLE_FILL
    return row + 1


def _write_header_table(ws, datos_causa, reajuste_tipo, start_row):
    row = _write_title(ws, "Datos de la causa", start_row)
    header_fill = PatternFill("solid", fgColor=_rgb_hex(config.COLOR_PJUD))

    for label, value in datos_causa:
        if str(label).lower() == "iniciales":
            continue
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = header_fill
        if str(label) == "Pension Final Arrastre":
            ws.cell(row=row, column=2).value = _currency_to_number(value)
            ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
        row += 1

    ws.cell(row=row, column=1, value="Tipo de Reajustabilidad")
    ws.cell(row=row, column=2, value=reajuste_tipo)
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = header_fill
    _style_range(ws, start_row + 1, row, 1, 2)
    return row + 2


def _write_period_table(ws, items_tabla, reajuste_tipo, start_row, cese_alimentos):
    if cese_alimentos:
        return start_row, None

    headers = _headers_for_reajuste(reajuste_tipo)
    row = _write_title(ws, "Detalle de calculo", start_row)
    header_row = row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    first_data_row = header_row + 1
    for item in items_tabla:
        if item and str(item[0]) == "TOTALES":
            continue
        for col, value in enumerate(item, start=1):
            _set_cell(ws.cell(row=row + 1, column=col), value, headers[col - 1])
        row += 1

    total_row = row + 1
    total_col = len(headers)
    ws.cell(row=total_row, column=1, value="TOTALES")
    ws.cell(row=total_row, column=total_col, value=f"=SUM({get_column_letter(total_col)}{first_data_row}:{get_column_letter(total_col)}{total_row - 1})")
    ws.cell(row=total_row, column=total_col).number_format = CURRENCY_FORMAT

    for col in range(1, total_col + 1):
        ws.cell(row=total_row, column=col).font = Font(bold=True)
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL

    _style_range(ws, header_row, total_row, 1, total_col)
    return total_row + 2, f"{get_column_letter(total_col)}{total_row}"


def _write_summary(ws, resumen, cartolas_data, start_row, period_total_ref, referencia_deuda_anterior, cese_alimentos):
    row = _write_title(ws, "Resumen editable", start_row)
    header_row = row
    headers = ["Tipo", "Descripcion", "Monto"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    row += 1

    def write_summary_row(tipo, descripcion, monto, fill=None, bold=False):
        nonlocal row
        ws.cell(row=row, column=1, value=tipo)
        ws.cell(row=row, column=2, value=descripcion)
        ws.cell(row=row, column=3, value=monto)
        ws.cell(row=row, column=3).number_format = CURRENCY_FORMAT
        if fill or bold:
            for col_index in range(1, 4):
                cell = ws.cell(row=row, column=col_index)
                if fill:
                    cell.fill = fill
                if bold:
                    cell.font = Font(bold=True)
        written_row = row
        row += 1
        return written_row

    def sum_formula(row_numbers):
        if not row_numbers:
            return 0
        return "=" + "+".join(f"C{row_number}" for row_number in row_numbers)

    ajustes = resumen.get("ajustes_manuales", [])
    charge_rows = []
    abono_rows = []

    if not cese_alimentos:
        charge_rows.append(
            write_summary_row(
                "Cargo",
                "Subtotal devengado periodo actual",
                f"={period_total_ref}" if period_total_ref else resumen.get("cargo_actual", 0),
            )
        )

    if resumen.get("deuda_anterior", 0) != 0:
        label = "Monto liquidacion anterior"
        if referencia_deuda_anterior:
            label += f" {referencia_deuda_anterior}"
        charge_rows.append(write_summary_row("Cargo", label, resumen["deuda_anterior"]))

    for ajuste in ajustes:
        if ajuste.get("tipo") != "Cargo":
            continue
        desc = ajuste.get("desc") or "Ajuste manual"
        fecha = ajuste.get("fecha")
        if fecha:
            desc = f"{desc} ({fecha})"
        charge_rows.append(write_summary_row("Cargo", desc, int(ajuste.get("monto", 0) or 0)))

    subtotal_cargos_row = write_summary_row(
        "Subtotal",
        "Subtotal pensiones devengadas",
        sum_formula(charge_rows),
        fill=TOTAL_FILL,
        bold=True,
    )

    if not cartolas_data:
        abono_rows.append(write_summary_row("Abono", "Abonos Cartola", 0))

    for index, cartola in enumerate(cartolas_data, start=1):
        lav = cartola.get("lav_number", "N/A")
        period = cartola.get("period", "No detectado")
        label = f"Cartola {index} - LAV N {lav}"
        if period and period != "No detectado":
            label += f" ({period})"
        abono_rows.append(write_summary_row("Abono", label, -abs(int(cartola.get("total_abonos", 0) or 0))))

    for ajuste in ajustes:
        if ajuste.get("tipo") != "Abono":
            continue
        desc = ajuste.get("desc") or "Ajuste manual"
        fecha = ajuste.get("fecha")
        if fecha:
            desc = f"{desc} ({fecha})"
        abono_rows.append(write_summary_row("Abono", desc, -abs(int(ajuste.get("monto", 0) or 0))))

    subtotal_abonos_row = write_summary_row(
        "Subtotal",
        "Subtotal abonos y descuentos",
        sum_formula(abono_rows),
        fill=TOTAL_FILL,
        bold=True,
    )

    write_summary_row(
        "Total",
        "Saldo final",
        f"=C{subtotal_cargos_row}+C{subtotal_abonos_row}",
        fill=TOTAL_FILL,
        bold=True,
    )

    _style_range(ws, header_row, row - 1, 1, 3)
    return row + 2


def _write_observations(ws, observaciones_finales, start_row):
    if not observaciones_finales:
        return start_row
    row = _write_title(ws, "Observaciones", start_row)
    ws.cell(row=row, column=1, value=observaciones_finales)
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=6)
    ws.cell(row=row, column=1).alignment = Alignment(vertical="top", wrap_text=True)
    return row + 4


def _write_cartola_sheet(wb, cartola, index, nombre_tribunal):
    ws = wb.create_sheet(title=f"Cartola {index}"[:31])
    row = _write_title(ws, f"Informe de busqueda de pensiones - Cartola {index}", 1)
    metadata = [
        ("RUT Titular", cartola.get("rut_dte", "No detectado")),
        ("Tribunal", nombre_tribunal),
        ("Cuenta LAV", cartola.get("lav_number", "N/A")),
        ("Periodo", cartola.get("period", "No detectado")),
        ("Total abonos", int(cartola.get("total_abonos", 0) or 0)),
    ]
    for label, value in metadata:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        if label == "Total abonos":
            ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
        row += 1

    row += 1
    headers = ["Fecha Mvto.", "Movimiento", "Monto"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    row += 1

    movimientos = cartola.get("movimientos", [])
    if not movimientos:
        ws.cell(row=row, column=1, value="No se encontraron movimientos para esta cartola.")
    else:
        for mov in movimientos:
            fecha = mov[0] if len(mov) > 0 else ""
            desc = mov[1] if len(mov) > 1 else ""
            monto = mov[2] if len(mov) > 2 else 0
            ws.cell(row=row, column=1, value=fecha)
            ws.cell(row=row, column=2, value=desc)
            ws.cell(row=row, column=3, value=int(monto or 0))
            ws.cell(row=row, column=3).number_format = CURRENCY_FORMAT
            row += 1

    _style_range(ws, 2, max(row - 1, 2), 1, 3)
    _autosize_columns(ws)
    ws.freeze_panes = "A8"


def generar_excel(datos_causa, items_tabla, resumen, ciudad, reajuste_tipo, cartolas_data=None, referencia_deuda_anterior="", observaciones_finales="", external_pdf_path=None, cese_alimentos=False, output_dir=None):
    cartolas_data = cartolas_data or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidacion"

    ws.cell(row=1, column=1, value="LIQUIDACION")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generado: {datetime.now():%d/%m/%Y %H:%M}")

    row = _write_header_table(ws, datos_causa, reajuste_tipo, 4)
    row, period_total_ref = _write_period_table(ws, items_tabla, reajuste_tipo, row, cese_alimentos)
    row = _write_summary(
        ws,
        resumen,
        cartolas_data,
        row,
        period_total_ref,
        referencia_deuda_anterior,
        cese_alimentos,
    )
    _write_observations(ws, observaciones_finales, row)

    _autosize_columns(ws)
    ws.freeze_panes = "A5"

    nombre_tribunal = datos_causa[0][1] if datos_causa else ""
    for index, cartola in enumerate(cartolas_data, start=1):
        _write_cartola_sheet(wb, cartola, index, nombre_tribunal)

    rit_causa = datos_causa[1][1] if len(datos_causa) > 1 else "SITFA"
    output_base_dir = output_dir or os.getcwd()
    os.makedirs(output_base_dir, exist_ok=True)
    output_path = os.path.join(output_base_dir, f"Liquidacion_{_safe_filename(rit_causa)}.xlsx")
    wb.save(output_path)
    return output_path
